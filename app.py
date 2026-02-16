from flask import Flask, render_template, request, redirect, url_for, g, send_file, session, flash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
from datetime import datetime
import os
import csv
import io
from openpyxl import load_workbook, Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from functools import wraps
import re

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'entregas.db')

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
    return db

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Check if old table structure exists
    cursor.execute("PRAGMA table_info(entregas)")
    columns = [column[1] for column in cursor.fetchall()]
    
    if columns and 'numero_serie' in columns:
        # Old structure, need to migrate
        conn.execute('''
            CREATE TABLE entregas_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                situm TEXT,
                usuario TEXT,
                imei TEXT,
                telefono TEXT,
                notas_telefono TEXT,
                tipo TEXT,
                timestamp TEXT
            )
        ''')
        conn.execute('''
            INSERT INTO entregas_new (situm, usuario, imei, telefono, notas_telefono, tipo, timestamp)
            SELECT '', usuario, imei, telefono, modelo, tipo, timestamp FROM entregas
        ''')
        conn.execute('DROP TABLE entregas')
        conn.execute('ALTER TABLE entregas_new RENAME TO entregas')
    elif not columns:
        # New table
        conn.execute('''
            CREATE TABLE entregas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                situm TEXT,
                usuario TEXT,
                imei TEXT,
                telefono TEXT,
                notas_telefono TEXT,
                tipo TEXT,
                timestamp TEXT
            )
        ''')
    
    # Create computers table if not exists
    cursor.execute("PRAGMA table_info(computers)")
    comp_columns = [column[1] for column in cursor.fetchall()]
    if not comp_columns:
        conn.execute('''
            CREATE TABLE computers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                hostname TEXT,
                numero_serie TEXT,
                apellidos_nombre TEXT,
                notas TEXT,
                tipo TEXT,
                usuario TEXT,
                timestamp TEXT,
                proyecto TEXT DEFAULT 'Mitie'
            )
        ''')
    else:
        if 'proyecto' not in comp_columns:
            conn.execute("ALTER TABLE computers ADD COLUMN proyecto TEXT DEFAULT 'Mitie'")
    
    # Create incidencias table if not exists
    cursor.execute("PRAGMA table_info(incidencias)")
    if not cursor.fetchall():
        conn.execute('''
            CREATE TABLE incidencias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                imei TEXT,
                usuario TEXT,
                telefono TEXT,
                notas TEXT,
                archivo_nombre TEXT,
                archivo_contenido BLOB,
                timestamp TEXT
            )
        ''')
    
    # Create usuarios table if not exists
    cursor.execute("PRAGMA table_info(usuarios)")
    user_columns = [column[1] for column in cursor.fetchall()]
    if not user_columns:
        conn.execute('''
            CREATE TABLE usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                rol TEXT NOT NULL,
                activo INTEGER DEFAULT 1,
                fecha_creacion TEXT,
                notas TEXT
            )
        ''')
        # Crear usuario admin por defecto
        admin_password = generate_password_hash('admin123')
        conn.execute('INSERT INTO usuarios (username, password, rol, activo, fecha_creacion) VALUES (?, ?, ?, ?, ?)',
                    ('admin', admin_password, 'admin', 1, datetime.utcnow().isoformat()))
    else:
        if 'notas' not in user_columns:
            conn.execute('ALTER TABLE usuarios ADD COLUMN notas TEXT')
    
    # Create usuarios_gtd_sgpmr table if not exists
    cursor.execute("PRAGMA table_info(usuarios_gtd_sgpmr)")
    if not cursor.fetchall():
        conn.execute('''
            CREATE TABLE usuarios_gtd_sgpmr (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                usuario_gtd TEXT,
                usuario_sgpmr TEXT,
                nombre_apellidos TEXT NOT NULL,
                correo_electronico TEXT,
                dni_nie TEXT,
                fecha_creacion TEXT
            )
        ''')
    
    # Create inventario_telefonos table if not exists
    cursor.execute("PRAGMA table_info(inventario_telefonos)")
    if not cursor.fetchall():
        conn.execute('''
            CREATE TABLE inventario_telefonos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                imei TEXT NOT NULL,
                numero_serie TEXT,
                modelo TEXT,
                telefono_asociado TEXT,
                fecha_creacion TEXT
            )
        ''')
    
    conn.commit()
    conn.close()

app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = 'your-secret-key-change-this-in-production'

# Configurar Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión'

# Rol y permisos basados en rol
ROLES_PERMISOS = {
    'admin': ['crear_usuario', 'eliminar_usuario', 'cambiar_rol', 'registrar', 'borrar_registros', 'ver_historico', 'ver_incidencias', 'administracion'],
    'operator': ['registrar', 'borrar_registros', 'ver_historico', 'ver_incidencias'],
    'viewer': ['ver_historico', 'ver_incidencias']
}

class User(UserMixin):
    def __init__(self, id, username, rol):
        self.id = id
        self.username = username
        self.rol = rol
    
    def tiene_permiso(self, permiso):
        return permiso in ROLES_PERMISOS.get(self.rol, [])

@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT id, username, rol FROM usuarios WHERE id = ? AND activo = 1', (user_id,))
    usuario = cursor.fetchone()
    if usuario:
        return User(usuario[0], usuario[1], usuario[2])
    return None

def require_permission(permiso):
    """Decorador para requerir un permiso específico"""
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated_function(*args, **kwargs):
            if not current_user.tiene_permiso(permiso):
                flash('No tienes permisos para realizar esta acción', 'error')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

init_db()

def format_phone(phone):
    """Normaliza teléfono: devuelve solo dígitos.

    - Si el campo está vacío devuelve cadena vacía.
    - Si contiene exactamente 9 dígitos devuelve esos 9 dígitos.
    - Si contiene otro formato devuelve None (indica inválido).
    No añade prefijo '34'.
    """
    if not phone:
        return ''
    phone = str(phone).strip()
    # Quitar cualquier carácter que no sea dígito
    digits = re.sub(r'\D', '', phone)
    # Si el usuario metió texto sin dígitos, considerar inválido
    if digits == '':
        return None

    # Caso ya en formato nacional (9 dígitos)
    if len(digits) == 9:
        return digits

    # Prefijo internacional +34 -> después de eliminar no-dígitos queda '34XXXXXXXXX'
    if digits.startswith('34') and len(digits) == 11:
        return digits[2:]

    # Prefijo internacional en forma 0034
    if digits.startswith('0034') and len(digits) == 13:
        return digits[4:]

    # Número con 0 inicial (ej. 0600123456) -> quitar 0 si produce 9 dígitos
    if digits.startswith('0') and len(digits) == 10:
        return digits[1:]

    # Si no coincide con ningún patrón válido, considerarlo inválido
    return None

def is_mitie_email(addr: str) -> bool:
    if not addr:
        return False
    addr = addr.strip()
    # allow case-insensitive mitie.es domain
    return re.match(r'^[A-Za-z0-9._%+-]+@mitie\.es$', addr, re.IGNORECASE) is not None

def is_valid_imei(imei: str) -> bool:
    if not imei:
        return False
    imei = imei.strip()
    return re.match(r'^\d{15}$', imei) is not None

def generate_entrega_pdf(situm, usuario, imei, telefono, notas, timestamp):
    """Genera PDF de entrega con datos y comunicación de geolocalización"""
    # Texto de comunicación
    texto_comunicacion = """Mitie Facilities Services, S.A. C/ Juan Ignacio Luca de Tena, 8 - 1° 28027 Madrid España

COMUNICACIÓN

Para mejorar la organización del servicio, evitar incidentes con los materiales de la Empresa, conseguir los estándares de calidad y mejorar la productividad que requiere nuestro servicio, a través de la entrega del dispositivo móvil (PDA) propiedad de la Empresa y de los equipos de trabajo que pone la Empresa a su disposición, se inicia la implantación de un sistema de geolocalización de estos dispositivos (PDA) y herramientas como sillas de ruedas cuyo uso se encuentra limitado a los fines anteriores y exclusivamente durante la prestación de SU jornada de trabajo en el Centro del Aeropuerto.

Para su conocimiento, se ha comunicado con anterioridad a la Representación Legal de las Personas de este Servicio PMR del Centro del Aeropuerto y con carácter previo a su implantación a través de diversas comunicaciones, principalmente efectuadas el 26 de septiembre, 11 de octubre y el pasado 28 de octubre del año en curso, con información detallada del sistema.

Le recordamos que el uso de los equipos y particularmente dispositivos móviles (PDA) entregados por la Empresa están destinados solo y exclusivamente para desempeño de su labor en el Servicio en el Centro del Aeropuerto, quedando el uso exclusivamente limitado a la prestación de sus servicios dentro de su jornada de trabajo, no estando permitido su utilización fuera de esta.

Por consiguiente:

1. La implantación no comporta modificación de sistema organizativo.

2. Se utiliza en herramientas y dispositivos móviles propiedad de la Empresa, protegiéndose la intimidad de todas las personas trabajadoras del servicio. Se trata por tanto de una geolocalización admitida, sobre Medios propiedad de la Empresa acordes al art. 20 bis ET.

3. Los datos relativos al posicionamiento geográfico que proporciona el uso del sistema de geolocalización, no constituyen datos de carácter personal salvo los ya recogido en la entrega del equipo, ya que se centran en medios y herramientas de la Empresa, si bien, garantizamos siempre y en todo caso que estos datos serán tratados de conformidad con lo dispuesto en el Reglamento (UE) 2016/679, de 27 de abril (GDPR), y la Ley Orgánica 3/2018, de 5 de diciembre (LOPDGDD), con el apoyo del Delegado de Protección de Datos (dpd@acoran.es D. Mario García.).

4. Este sistema cumple con todos los requisitos a nivel legal, respetando en todo momento los principios de proporcionalidad, necesidad e idoneidad.

5. La Geolocalización, le recordamos, solo se realiza cuándo el equipo está operativo, durante su jornada de trabajo y en las zonas públicas del entorno Aeroportuario, respetando los periodos de descanso como refrigerio.
"""
    
    # Crear PDF en memoria
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12,
        alignment=1  # Centrado
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        spaceAfter=6
    )
    
    # Contenido del PDF
    elements = []
    
    # Añadir logo si existe (coloca tu archivo en static/mitie_logo.png)
    logo_path = os.path.join(BASE_DIR, 'static', 'mitie_logo.png')
    if os.path.exists(logo_path):
        try:
            # Leer tamaño real y escalar manteniendo proporción
            img_reader = ImageReader(logo_path)
            img_w, img_h = img_reader.getSize()  # en píxeles

            # Ancho objetivo por defecto (2 pulgadas)
            desired_width = 2 * inch
            # Calcular altura manteniendo proporción
            desired_height = desired_width * (float(img_h) / float(img_w)) if img_w else desired_width * 0.5

            # Calcular altura máxima disponible (25% de la altura útil de la página)
            page_height = doc.pagesize[1]
            usable_height = page_height - doc.topMargin - doc.bottomMargin
            max_height = usable_height * 0.25

            if desired_height > max_height:
                scale = max_height / desired_height
                desired_width = desired_width * scale
                desired_height = max_height

            logo = Image(logo_path, width=desired_width, height=desired_height)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.15*inch))
        except Exception:
            # Si no se puede cargar o medir el logo, continuar sin él
            pass

    # Encabezado
    elements.append(Paragraph("REGISTRO DE ENTREGA DE DISPOSITIVO", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Tabla de datos
    data = [
        ['CAMPO', 'VALOR'],
        ['Situm', situm or ''],
        ['Usuario', usuario or ''],
        ['IMEI', imei or ''],
        ['Teléfono', telefono or ''],
        ['Notas', notas or ''],
        ['Fecha', timestamp[:10] if timestamp else '']
    ]
    
    table = Table(data, colWidths=[1.5*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.HexColor('#ffffff')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Texto de comunicación
    elements.append(Paragraph("COMUNICACIÓN SOBRE SISTEMA DE GEOLOCALIZACIÓN", title_style))
    elements.append(Spacer(1, 0.1*inch))
    elements.append(Paragraph(texto_comunicacion, normal_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Firma
    elements.append(Paragraph("Recibí por: ________________________", normal_style))
    elements.append(Spacer(1, 0.05*inch))
    elements.append(Paragraph("Nombre y Apellidos: ________________________", normal_style))
    elements.append(Spacer(1, 0.05*inch))
    elements.append(Paragraph("D.N.I: ________________________", normal_style))
    elements.append(Spacer(1, 0.05*inch))
    elements.append(Paragraph("Fecha: ________________________", normal_style))
    
    # Construir PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    # Guardar copia en servidor
    pdf_dir = os.path.join(BASE_DIR, 'pdfs', 'entregas')
    os.makedirs(pdf_dir, exist_ok=True)
    pdf_filename = f"entrega_{imei}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf_path = os.path.join(pdf_dir, pdf_filename)
    with open(pdf_path, 'wb') as f:
        f.write(pdf_buffer.getvalue())
    
    # Retornar para descargar
    pdf_buffer.seek(0)
    return pdf_buffer, pdf_filename

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

@app.route('/')
def index():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    return render_template('index.html')

@app.route('/entrega_moviles', methods=['GET'])
def entrega_moviles():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    return render_template('entrega_moviles.html')

@app.route('/recepcion_moviles', methods=['GET'])
def recepcion_moviles():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    return render_template('recepcion_moviles.html')

@app.route('/incidencias_moviles', methods=['GET'])
def incidencias_moviles():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    return render_template('incidencias_moviles.html')

@app.route('/entrega_computer', methods=['GET', 'POST'])
def entrega_computer():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        hostname = request.form.get('hostname', '').strip()
        numero_serie = request.form.get('numero_serie', '').strip()
        apellidos_nombre = request.form.get('apellidos_nombre', '').strip()
        notas = request.form.get('notas', '').strip()
        usuario = current_user.username
        timestamp = datetime.now().isoformat()
        
        if not hostname:
            flash('Hostname es requerido', 'error')
            return render_template('entrega_computer.html')
        
        db = get_db()
        db.execute(
            'INSERT INTO computers (hostname, numero_serie, apellidos_nombre, notas, tipo, usuario, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (hostname, numero_serie, apellidos_nombre, notas, 'Entrega', usuario, timestamp)
        )
        db.commit()
        flash('Computer entregado correctamente', 'success')
        return redirect(url_for('index'))
    
    return render_template('entrega_computer.html')

@app.route('/recepcion_computer', methods=['GET', 'POST'])
def recepcion_computer():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        hostname = request.form.get('hostname', '').strip()
        numero_serie = request.form.get('numero_serie', '').strip()
        apellidos_nombre = request.form.get('apellidos_nombre', '').strip()
        notas = request.form.get('notas', '').strip()
        usuario = current_user.username
        timestamp = datetime.now().isoformat()
        
        if not hostname:
            flash('Hostname es requerido', 'error')
            return render_template('recepcion_computer.html')
        
        db = get_db()
        db.execute(
            'INSERT INTO computers (hostname, numero_serie, apellidos_nombre, notas, tipo, usuario, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (hostname, numero_serie, apellidos_nombre, notas, 'Recepción', usuario, timestamp)
        )
        db.commit()
        flash('Computer recibido correctamente', 'success')
        return redirect(url_for('index'))
    
    return render_template('recepcion_computer.html')

@app.route('/incidencias_computer', methods=['GET', 'POST'])
def incidencias_computer():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        hostname = request.form.get('hostname', '').strip()
        numero_serie = request.form.get('numero_serie', '').strip()
        apellidos_nombre = request.form.get('apellidos_nombre', '').strip()
        notas = request.form.get('notas', '').strip()
        usuario = current_user.username
        timestamp = datetime.now().isoformat()
        
        if not hostname:
            flash('Hostname es requerido', 'error')
            return render_template('incidencias_computer.html')
        
        db = get_db()
        db.execute(
            'INSERT INTO computers (hostname, numero_serie, apellidos_nombre, notas, tipo, usuario, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (hostname, numero_serie, apellidos_nombre, notas, 'Incidencia', usuario, timestamp)
        )
        db.commit()
        flash('Incidencia registrada correctamente', 'success')
        return redirect(url_for('index'))
    
    return render_template('incidencias_computer.html')

@app.route('/Entrada_computer_aena', methods=['GET', 'POST'])
def Entrada_computer_aena():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        hostname = request.form.get('hostname', '').strip()
        numero_serie = request.form.get('numero_serie', '').strip()
        apellidos_nombre = request.form.get('apellidos_nombre', '').strip()
        notas = request.form.get('notas', '').strip()
        usuario = current_user.username
        timestamp = datetime.now().isoformat()
        
        if not hostname:
            flash('Hostname es requerido', 'error')
            return render_template('entrega_computer_aena.html')
        
        db = get_db()
        db.execute(
            'INSERT INTO computers (hostname, numero_serie, apellidos_nombre, notas, tipo, usuario, timestamp, proyecto) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (hostname, numero_serie, apellidos_nombre, notas, 'Entrega', usuario, timestamp, 'AENA')
        )
        db.commit()
        flash('Computer AENA entregado correctamente', 'success')
        return redirect(url_for('index'))
    
    return render_template('entrega_computer_aena.html')

@app.route('/incidencias_computer_aena', methods=['GET', 'POST'])
def incidencias_computer_aena():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        hostname = request.form.get('hostname', '').strip()
        numero_serie = request.form.get('numero_serie', '').strip()
        apellidos_nombre = request.form.get('apellidos_nombre', '').strip()
        notas = request.form.get('notas', '').strip()
        usuario = current_user.username
        timestamp = datetime.now().isoformat()
        
        if not hostname:
            flash('Hostname es requerido', 'error')
            return render_template('incidencias_computer_aena.html')
        
        db = get_db()
        db.execute(
            'INSERT INTO computers (hostname, numero_serie, apellidos_nombre, notas, tipo, usuario, timestamp, proyecto) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (hostname, numero_serie, apellidos_nombre, notas, 'Incidencia', usuario, timestamp, 'AENA')
        )
        db.commit()
        flash('Incidencia AENA registrada correctamente', 'success')
        return redirect(url_for('index'))
    
    return render_template('incidencias_computer_aena.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('Usuario y contraseña requeridos', 'error')
            return redirect(url_for('login'))
        
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT id, username, password, rol, activo FROM usuarios WHERE username = ?', (username,))
        usuario = cursor.fetchone()
        
        if usuario and usuario[4] == 1 and check_password_hash(usuario[2], password):
            user = User(usuario[0], usuario[1], usuario[3])
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Sesión cerrada correctamente', 'success')
    return redirect(url_for('login'))

@app.route('/verificar_password_borrado', methods=['POST'])
@login_required
def verificar_password_borrado():
    data = request.get_json()
    password = data.get('password')
    
    if not password:
        return {'success': False, 'message': 'Contraseña requerida'}, 400
        
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT password FROM usuarios WHERE id = ?', (current_user.id,))
    row = cursor.fetchone()
    
    if row and check_password_hash(row[0], password):
        return {'success': True}
    else:
        return {'success': False, 'message': 'Contraseña incorrecta'}

@app.route('/administracion')
@require_permission('administracion')
def administracion():
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT id, username, rol, activo, fecha_creacion FROM usuarios ORDER BY fecha_creacion DESC')
    usuarios = cursor.fetchall()
    return render_template('administracion.html', usuarios=usuarios, roles=ROLES_PERMISOS.keys())

@app.route('/usuarios/crear', methods=['GET', 'POST'])
@require_permission('crear_usuario')
def crear_usuario():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        rol = request.form.get('rol', 'viewer')
        
        # Permitimos cualquier contraseña no vacía proporcionada por el administrador.
        if not username or password is None or password == '':
            flash('Usuario y contraseña requeridos', 'error')
            return redirect(url_for('crear_usuario'))
        
        if rol not in ROLES_PERMISOS:
            flash('Rol inválido', 'error')
            return redirect(url_for('crear_usuario'))
        
        db = get_db()
        try:
            hashed_password = generate_password_hash(password)
            db.execute('INSERT INTO usuarios (username, password, rol, activo, fecha_creacion) VALUES (?, ?, ?, ?, ?)',
                      (username, hashed_password, rol, 1, datetime.utcnow().isoformat()))
            db.commit()
            flash(f'Usuario {username} creado correctamente', 'success')
            return redirect(url_for('administracion'))
        except sqlite3.IntegrityError:
            flash(f'El usuario {username} ya existe', 'error')
            return redirect(url_for('crear_usuario'))
    
    return render_template('crear_usuario.html', roles=ROLES_PERMISOS.keys())

@app.route('/usuarios/<int:usuario_id>/editar', methods=['GET', 'POST'])
@require_permission('cambiar_rol')
def editar_usuario(usuario_id):
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute('SELECT id, username, rol, activo FROM usuarios WHERE id = ?', (usuario_id,))
    usuario = cursor.fetchone()
    
    if not usuario:
        flash('Usuario no encontrado', 'error')
        return redirect(url_for('administracion'))
    
    if request.method == 'POST':
        rol = request.form.get('rol', 'viewer')
        activo = request.form.get('activo', '0')
        
        if rol not in ROLES_PERMISOS:
            flash('Rol inválido', 'error')
            return redirect(url_for('editar_usuario', usuario_id=usuario_id))
        
        activo_int = 1 if activo == '1' else 0
        db.execute('UPDATE usuarios SET rol = ?, activo = ? WHERE id = ?', (rol, activo_int, usuario_id))
        db.commit()
        flash(f'Usuario actualizado correctamente', 'success')
        return redirect(url_for('administracion'))
    
    return render_template(
        'editar_usuario.html',
        usuario=usuario,
        roles=ROLES_PERMISOS.keys(),
        roles_permisos=ROLES_PERMISOS,
    )


@app.route('/usuarios/<int:usuario_id>/cambiar_contrasena', methods=['GET', 'POST'])
@require_permission('cambiar_rol')
def cambiar_contrasena_usuario(usuario_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT id, username FROM usuarios WHERE id = ?', (usuario_id,))
    usuario = cursor.fetchone()
    if not usuario:
        flash('Usuario no encontrado', 'error')
        return redirect(url_for('administracion'))

    if request.method == 'POST':
        new = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')

        if not new or not confirm:
            flash('Todos los campos son requeridos', 'error')
            return redirect(url_for('cambiar_contrasena_usuario', usuario_id=usuario_id))

        if new != confirm:
            flash('La nueva contraseña y la confirmación no coinciden', 'error')
            return redirect(url_for('cambiar_contrasena_usuario', usuario_id=usuario_id))

        new_hash = generate_password_hash(new)
        db.execute('UPDATE usuarios SET password = ? WHERE id = ?', (new_hash, usuario_id))
        db.commit()
        flash('Contraseña actualizada correctamente', 'success')
        return redirect(url_for('administracion'))

    # usuario puede ser sqlite3.Row; pasar username para mostrar
    username = usuario['username'] if hasattr(usuario, 'keys') else usuario[1]
    return render_template('admin_cambiar_contrasena.html', usuario_id=usuario_id, username=username)

@app.route('/usuarios/<int:usuario_id>/eliminar', methods=['POST'])
@require_permission('eliminar_usuario')
def eliminar_usuario(usuario_id):
    if usuario_id == current_user.id:
        flash('No puedes eliminar tu propia cuenta', 'error')
        return redirect(url_for('administracion'))
    
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT username FROM usuarios WHERE id = ?', (usuario_id,))
    usuario = cursor.fetchone()
    
    if not usuario:
        flash('Usuario no encontrado', 'error')
        return redirect(url_for('administracion'))
    
    db.execute('DELETE FROM usuarios WHERE id = ?', (usuario_id,))
    db.commit()
    flash(f'Usuario eliminado correctamente', 'success')
    return redirect(url_for('administracion'))

@app.route('/entrega', methods=['POST'])
@require_permission('registrar')
def entrega():
    situm = request.form.get('situm', '').strip()
    usuario = request.form.get('usuario', '').strip()
    imei = request.form.get('imei', '').strip()
    raw_telefono = request.form.get('telefono', '').strip()
    telefono = format_phone(raw_telefono)
    notas_telefono = request.form.get('notas_telefono', '').strip()
    tipo = 'entrega'
    timestamp = datetime.utcnow().isoformat()
    db = get_db()

    # Validación server-side del campo situm (dominio @mitie.es)
    if situm and not is_mitie_email(situm):
        flash('El campo Situm debe ser un email con dominio @mitie.es', 'error')
        return redirect(url_for('index'))

    # Normalizar IMEI (quitar todo lo que no sean dígitos) y validación
    imei_digits = re.sub(r'\D', '', imei or '')
    if imei and not is_valid_imei(imei_digits):
        flash('IMEI inválido — debe contener exactamente 15 dígitos', 'error')
        return redirect(url_for('index'))
    imei = imei_digits

    # Validar teléfono (opcional): si se ha rellenado debe ser 9 dígitos numéricos
    if raw_telefono and telefono is None:
        flash('Teléfono inválido — debe ser numérico de 9 dígitos sin prefijo 34', 'error')
        return redirect(url_for('index'))

    # Comprobar si el IMEI ya está 'entregado' (es decir, no ha sido recepcionado aún)
    if imei:
        cursor = db.cursor()
        cursor.execute('SELECT tipo FROM entregas WHERE imei = ? ORDER BY timestamp DESC LIMIT 1', (imei,))
        last = cursor.fetchone()
        if last and last['tipo'] == 'entrega':
            flash(f'No se puede registrar la entrega: el dispositivo con IMEI {imei} no ha sido recepcionado aún.', 'error')
            return redirect(url_for('index'))

    db.execute('INSERT INTO entregas (situm, usuario, imei, telefono, notas_telefono, tipo, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)',
               (situm, usuario, imei, telefono, notas_telefono, tipo, timestamp))
    db.commit()
    
    # Generar PDF
    pdf_buffer, pdf_filename = generate_entrega_pdf(situm, usuario, imei, telefono, notas_telefono, timestamp)
    
    # Devolver PDF para descargar
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=pdf_filename,
        mimetype='application/pdf'
    )

@app.route('/recepcion', methods=['POST'])
@require_permission('registrar')
def recepcion():
    situm = request.form.get('situm', '').strip()
    usuario = request.form.get('usuario', '').strip()
    imei = request.form.get('imei', '').strip()
    raw_telefono = request.form.get('telefono', '').strip()
    telefono = format_phone(raw_telefono)
    notas_telefono = request.form.get('notas_telefono', '').strip()
    tipo = 'recepcion'
    timestamp = datetime.utcnow().isoformat()
    db = get_db()

    # Validación server-side del campo situm (dominio @mitie.es)
    if situm and not is_mitie_email(situm):
        flash('El campo Situm debe ser un email con dominio @mitie.es', 'error')
        return redirect(url_for('index'))

    # Normalizar IMEI (quitar todo lo que no sean dígitos) y validación
    imei_digits = re.sub(r'\D', '', imei or '')
    if imei and not is_valid_imei(imei_digits):
        flash('IMEI inválido — debe contener exactamente 15 dígitos', 'error')
        return redirect(url_for('index'))
    imei = imei_digits
    # Validar teléfono (opcional)
    if raw_telefono and telefono is None:
        flash('Teléfono inválido — debe ser numérico de 9 dígitos sin prefijo 34', 'error')
        return redirect(url_for('index'))
    db.execute('INSERT INTO entregas (situm, usuario, imei, telefono, notas_telefono, tipo, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)',
               (situm, usuario, imei, telefono, notas_telefono, tipo, timestamp))
    db.commit()
    return redirect(url_for('index'))

@app.route('/incidencia', methods=['POST'])
@require_permission('registrar')
def incidencia():
    usuario = request.form.get('usuario', '').strip()
    imei = request.form.get('imei', '').strip()
    raw_telefono = request.form.get('telefono', '').strip()
    telefono = format_phone(raw_telefono)
    notas = request.form.get('notas', '').strip()
    timestamp = datetime.utcnow().isoformat()

    # Normalizar IMEI (quitar todo lo que no sean dígitos) y validación
    imei_digits = re.sub(r'\D', '', imei or '')
    if imei and not is_valid_imei(imei_digits):
        flash('IMEI inválido — debe contener exactamente 15 dígitos', 'error')
        return redirect(url_for('index'))
    imei = imei_digits
    
    # Archivo opcional: si se sube se valida y se almacena, si no se guarda vacío
    archivo_nombre = None
    archivo_contenido = None
    if 'archivo' in request.files and request.files['archivo'].filename != '':
        archivo = request.files['archivo']
        # Validar tipo de archivo
        allowed_extensions = {'pdf', 'jpg', 'jpeg'}
        if not ('.' in archivo.filename and archivo.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return redirect(url_for('index') + '?error=Invalid file type')
        # Leer contenido del archivo
        archivo_contenido = archivo.read()
        archivo_nombre = archivo.filename
    
    # Guardar en BD
    db = get_db()
    # Validar teléfono (opcional)
    if raw_telefono and telefono is None:
        flash('Teléfono inválido — debe ser numérico de 9 dígitos sin prefijo 34', 'error')
        return redirect(url_for('index'))

    db.execute('INSERT INTO incidencias (imei, usuario, telefono, notas, archivo_nombre, archivo_contenido, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)',
               (imei, usuario, telefono, notas, archivo_nombre, archivo_contenido, timestamp))
    db.commit()
    return redirect(url_for('index'))

@app.route('/history/clear', methods=['POST'])
@require_permission('borrar_registros')
def clear_history():
    """Borrar todo el histórico con verificación de contraseña"""
    password = request.form.get('password', '').strip()
    
    # Verificar contraseña
    if password != '4j_6WbTT7scyicJcam':
        return redirect(url_for('history_entrega') + '?error=Invalid password')
    
    db = get_db()
    db.execute('DELETE FROM entregas')
    db.commit()
    return redirect(url_for('history_entrega'))


def _delete_selected_history(ids_param, tipos, redirect_endpoint):
    """Borrar registros seleccionados con verificación de contraseña."""
    password = request.form.get('password', '').strip()

    if password != '4j_6WbTT7scyicJcam':
        return redirect(url_for(redirect_endpoint) + '?error=Invalid password')

    if ids_param:
        id_list = [int(i) for i in ids_param.split(',') if i.strip().isdigit()]
        if id_list:
            db = get_db()
            placeholders = ','.join(['?'] * len(id_list))
            tipo_placeholders = ','.join(['?'] * len(tipos))
            query = (
                f'DELETE FROM entregas '
                f'WHERE id IN ({placeholders}) '
                f'AND LOWER(tipo) IN ({tipo_placeholders})'
            )
            db.execute(query, id_list + tipos)
            db.commit()

    return redirect(url_for(redirect_endpoint))


@app.route('/history/delete-selected', methods=['POST'])
@require_permission('borrar_registros')
def delete_selected():
    ids_param = request.form.get('ids', '').strip()
    return _delete_selected_history(ids_param, ['entrega', 'entregas'], 'history_entrega')


@app.route('/history_entrega/delete-selected', methods=['POST'])
@require_permission('borrar_registros')
def delete_selected_entrega():
    ids_param = request.form.get('ids', '').strip()
    return _delete_selected_history(ids_param, ['entrega', 'entregas'], 'history_entrega')


@app.route('/history_recepcion/delete-selected', methods=['POST'])
@require_permission('borrar_registros')
def delete_selected_recepcion():
    ids_param = request.form.get('ids', '').strip()
    return _delete_selected_history(ids_param, ['recepcion', 'recepción', 'recepciones'], 'history_recepcion')


@app.route('/history/export')
@require_permission('ver_historico')
def export_history():
    return redirect(url_for('export_history_entrega'))

@app.route('/history_entrega/export')
@require_permission('ver_historico')
def export_history_entrega():
    db = get_db()
    imei_search = request.args.get('imei', '').strip()
    usuario_search = request.args.get('usuario', '').strip()
    ids_param = request.args.get('ids', '').strip()

    # If specific ids provided, export only those
    if ids_param:
        id_list = [int(i) for i in ids_param.split(',') if i.strip().isdigit()]
        if not id_list:
            rows = []
        else:
            placeholders = ','.join(['?'] * len(id_list))
            query = f'SELECT id, tipo, situm, usuario, imei, telefono, notas_telefono, timestamp FROM entregas WHERE LOWER(tipo) IN ("entrega", "entregas") AND id IN ({placeholders}) ORDER BY timestamp DESC'
            cur = db.execute(query, id_list)
            rows = cur.fetchall()
    else:
        query = 'SELECT id, tipo, situm, usuario, imei, telefono, notas_telefono, timestamp FROM entregas WHERE LOWER(tipo) IN ("entrega", "entregas") AND 1=1'
        params = []
        if imei_search:
            query += ' AND imei LIKE ?'
            params.append(f'%{imei_search}%')
        if usuario_search:
            query += ' AND usuario LIKE ?'
            params.append(f'%{usuario_search}%')
        query += ' ORDER BY timestamp DESC'

        cur = db.execute(query, params)
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.append(['Situm', 'Usuario', 'IMEI', 'Teléfono', 'Notas de Teléfono', 'Fecha (UTC)'])
    for r in rows:
        ws.append([
            r['situm'] if r['situm'] else '',
            r['usuario'] if r['usuario'] else '',
            r['imei'] if r['imei'] else '',
            r['telefono'] if r['telefono'] else '',
            r['notas_telefono'] if r['notas_telefono'] else '',
            r['timestamp'] if r['timestamp'] else '',
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name='historico_entregas.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/history_recepcion/export')
@require_permission('ver_historico')
def export_history_recepcion():
    db = get_db()
    imei_search = request.args.get('imei', '').strip()
    usuario_search = request.args.get('usuario', '').strip()
    ids_param = request.args.get('ids', '').strip()

    # If specific ids provided, export only those
    if ids_param:
        id_list = [int(i) for i in ids_param.split(',') if i.strip().isdigit()]
        if not id_list:
            rows = []
        else:
            placeholders = ','.join(['?'] * len(id_list))
            query = f'SELECT id, tipo, situm, usuario, imei, telefono, notas_telefono, timestamp FROM entregas WHERE LOWER(tipo) IN ("recepción", "recepcion", "recepciones") AND id IN ({placeholders}) ORDER BY timestamp DESC'
            cur = db.execute(query, id_list)
            rows = cur.fetchall()
    else:
        query = 'SELECT id, tipo, situm, usuario, imei, telefono, notas_telefono, timestamp FROM entregas WHERE LOWER(tipo) IN ("recepción", "recepcion", "recepciones") AND 1=1'
        params = []
        if imei_search:
            query += ' AND imei LIKE ?'
            params.append(f'%{imei_search}%')
        if usuario_search:
            query += ' AND usuario LIKE ?'
            params.append(f'%{usuario_search}%')
        query += ' ORDER BY timestamp DESC'

        cur = db.execute(query, params)
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.append(['Situm', 'Usuario', 'IMEI', 'Teléfono', 'Notas de Teléfono', 'Fecha (UTC)'])
    for r in rows:
        ws.append([
            r['situm'] if r['situm'] else '',
            r['usuario'] if r['usuario'] else '',
            r['imei'] if r['imei'] else '',
            r['telefono'] if r['telefono'] else '',
            r['notas_telefono'] if r['notas_telefono'] else '',
            r['timestamp'] if r['timestamp'] else '',
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name='historico_recepciones.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/history')
@require_permission('ver_historico')
def history():
    # Redirigir a history_entrega por defecto
    return redirect('/history_entrega')

@app.route('/history_entrega')
@require_permission('ver_historico')
def history_entrega():
    db = get_db()
    imei_search = request.args.get('imei', '').strip()
    usuario_search = request.args.get('usuario', '').strip()
    fecha_inicio = request.args.get('fecha_inicio', '').strip()
    fecha_fin = request.args.get('fecha_fin', '').strip()
    
    query = 'SELECT * FROM entregas WHERE LOWER(tipo) IN ("entrega", "entregas") AND 1=1'
    params = []
    
    if imei_search:
        query += ' AND imei LIKE ?'
        params.append(f'%{imei_search}%')
    
    if usuario_search:
        query += ' AND usuario LIKE ?'
        params.append(f'%{usuario_search}%')
    
    if fecha_inicio:
        query += ' AND timestamp >= ?'
        params.append(f'{fecha_inicio}T00:00:00')
    
    if fecha_fin:
        query += ' AND timestamp <= ?'
        params.append(f'{fecha_fin}T23:59:59')
    
    query += ' ORDER BY timestamp DESC'
    
    cur = db.execute(query, params)
    rows = cur.fetchall()
    return render_template('history_entrega.html', rows=rows, imei_search=imei_search, usuario_search=usuario_search, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

@app.route('/history_recepcion')
@require_permission('ver_historico')
def history_recepcion():
    db = get_db()
    imei_search = request.args.get('imei', '').strip()
    usuario_search = request.args.get('usuario', '').strip()
    fecha_inicio = request.args.get('fecha_inicio', '').strip()
    fecha_fin = request.args.get('fecha_fin', '').strip()
    
    query = 'SELECT * FROM entregas WHERE LOWER(tipo) IN ("recepción", "recepcion", "recepciones") AND 1=1'
    params = []
    
    if imei_search:
        query += ' AND imei LIKE ?'
        params.append(f'%{imei_search}%')
    
    if usuario_search:
        query += ' AND usuario LIKE ?'
        params.append(f'%{usuario_search}%')
    
    if fecha_inicio:
        query += ' AND timestamp >= ?'
        params.append(f'{fecha_inicio}T00:00:00')
    
    if fecha_fin:
        query += ' AND timestamp <= ?'
        params.append(f'{fecha_fin}T23:59:59')
    
    query += ' ORDER BY timestamp DESC'
    
    cur = db.execute(query, params)
    rows = cur.fetchall()
    return render_template('history_recepcion.html', rows=rows, imei_search=imei_search, usuario_search=usuario_search, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)


@app.route('/history_computers_entrega')
@require_permission('ver_historico')
def history_computers_entrega():
    db = get_db()
    hostname_search = request.args.get('hostname', '').strip()
    sn_search = request.args.get('sn', '').strip()
    proyecto_filter = request.args.get('proyecto', '').strip()
    
    query = 'SELECT * FROM computers WHERE tipo = "Entrega" AND 1=1'
    params = []
    
    if hostname_search:
        query += ' AND hostname LIKE ?'
        params.append(f'%{hostname_search}%')
    if sn_search:
        query += ' AND numero_serie LIKE ?'
        params.append(f'%{sn_search}%')
    if proyecto_filter:
        query += ' AND proyecto = ?'
        params.append(proyecto_filter)
        
    query += ' ORDER BY timestamp DESC'
    rows = db.execute(query, params).fetchall()
    return render_template('history_computers.html', rows=rows, title="Histórico Entregas Computer", 
                           hostname_search=hostname_search, sn_search=sn_search, proyecto_filter=proyecto_filter,
                           tipo_actual="Entrega")

@app.route('/history_computers_recepcion')
@require_permission('ver_historico')
def history_computers_recepcion():
    db = get_db()
    hostname_search = request.args.get('hostname', '').strip()
    sn_search = request.args.get('sn', '').strip()
    proyecto_filter = request.args.get('proyecto', '').strip()
    
    query = 'SELECT * FROM computers WHERE tipo = "Recepción" AND 1=1'
    params = []
    
    if hostname_search:
        query += ' AND hostname LIKE ?'
        params.append(f'%{hostname_search}%')
    if sn_search:
        query += ' AND numero_serie LIKE ?'
        params.append(f'%{sn_search}%')
    if proyecto_filter:
        query += ' AND proyecto = ?'
        params.append(proyecto_filter)
        
    query += ' ORDER BY timestamp DESC'
    rows = db.execute(query, params).fetchall()
    return render_template('history_computers.html', rows=rows, title="Histórico Recepciones Computer", 
                           hostname_search=hostname_search, sn_search=sn_search, proyecto_filter=proyecto_filter,
                           tipo_actual="Recepción")

@app.route('/history_computers_incidencias')
@require_permission('ver_historico')
def history_computers_incidencias():
    db = get_db()
    hostname_search = request.args.get('hostname', '').strip()
    sn_search = request.args.get('sn', '').strip()
    proyecto_filter = request.args.get('proyecto', '').strip()
    
    query = 'SELECT * FROM computers WHERE tipo = "Incidencia" AND 1=1'
    params = []
    
    if hostname_search:
        query += ' AND hostname LIKE ?'
        params.append(f'%{hostname_search}%')
    if sn_search:
        query += ' AND numero_serie LIKE ?'
        params.append(f'%{sn_search}%')
    if proyecto_filter:
        query += ' AND proyecto = ?'
        params.append(proyecto_filter)
        
    query += ' ORDER BY timestamp DESC'
    rows = db.execute(query, params).fetchall()
    return render_template('history_computers.html', rows=rows, title="Histórico Incidencias Computer", 
                           hostname_search=hostname_search, sn_search=sn_search, proyecto_filter=proyecto_filter,
                           tipo_actual="Incidencia")


@app.route('/history_computers/export')
@require_permission('ver_historico')
def export_history_computers():
    db = get_db()
    hostname_search = request.args.get('hostname', '').strip()
    sn_search = request.args.get('sn', '').strip()
    proyecto_filter = request.args.get('proyecto', '').strip()
    tipo_filter = request.args.get('tipo', '').strip()

    query = 'SELECT proyecto, hostname, numero_serie, apellidos_nombre, notas, usuario, timestamp, tipo FROM computers WHERE 1=1'
    params = []

    if tipo_filter:
        query += ' AND tipo = ?'
        params.append(tipo_filter)
    if hostname_search:
        query += ' AND hostname LIKE ?'
        params.append(f'%{hostname_search}%')
    if sn_search:
        query += ' AND numero_serie LIKE ?'
        params.append(f'%{sn_search}%')
    if proyecto_filter:
        query += ' AND proyecto = ?'
        params.append(proyecto_filter)

    query += ' ORDER BY timestamp DESC'
    rows = db.execute(query, params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.append(['Proyecto', 'Hostname', 'S/N', 'Persona', 'Notas', 'Registrado por', 'Fecha', 'Tipo'])
    for r in rows:
        ws.append([
            r['proyecto'] or 'Mitie',
            r['hostname'] or '',
            r['numero_serie'] or '',
            r['apellidos_nombre'] or '',
            r['notas'] or '',
            r['usuario'] or '',
            r['timestamp'] or '',
            r['tipo'] or ''
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"historico_computers_{tipo_filter or 'todos'}.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/history_computers/delete-selected', methods=['POST'])
@require_permission('borrar_registros')
def delete_selected_computers():
    """Borrar registros de computers seleccionados con verificación de contraseña."""
    password = request.form.get('password', '').strip()
    ids_param = request.form.get('ids', '').strip()

    if password != '4j_6WbTT7scyicJcam':
        return redirect(request.referrer or url_for('index'))

    if ids_param:
        id_list = [int(i) for i in ids_param.split(',') if i.strip().isdigit()]
        if id_list:
            db = get_db()
            placeholders = ','.join(['?'] * len(id_list))
            query = f'DELETE FROM computers WHERE id IN ({placeholders})'
            db.execute(query, id_list)
            db.commit()

    return redirect(request.referrer or url_for('index'))


@app.route('/history_computers/<int:computer_id>/editar', methods=['GET', 'POST'])
@require_permission('administracion')
def editar_computer(computer_id):
    db = get_db()
    cur = db.execute('SELECT * FROM computers WHERE id = ?', (computer_id,))
    r = cur.fetchone()
    if not r:
        flash('Registro no encontrado', 'error')
        return redirect(request.referrer or url_for('index'))

    if request.method == 'GET':
        return render_template('edit_computer.html', r=r)

    # POST: procesar actualización
    proyecto = request.form.get('proyecto', 'Mitie').strip()
    hostname = request.form.get('hostname', '').strip()
    numero_serie = request.form.get('numero_serie', '').strip()
    apellidos_nombre = request.form.get('apellidos_nombre', '').strip()
    notas = request.form.get('notas', '').strip()
    tipo = request.form.get('tipo', 'Entrega').strip()

    db.execute('''
        UPDATE computers 
        SET proyecto = ?, hostname = ?, numero_serie = ?, apellidos_nombre = ?, notas = ?, tipo = ?
        WHERE id = ?
    ''', (proyecto, hostname, numero_serie, apellidos_nombre, notas, tipo, computer_id))
    db.commit()

    flash('Registro actualizado correctamente', 'success')
    
    # Redirigir según el tipo de registro actualizado
    if tipo == "Entrega":
        return redirect(url_for('history_computers_entrega'))
    elif tipo == "Recepción":
        return redirect(url_for('history_computers_recepcion'))
    else:
        return redirect(url_for('history_computers_incidencias'))


@app.route('/history_computers/import', methods=['GET', 'POST'])
@require_permission('registrar')
def import_computers():
    if request.method == 'GET':
        return render_template('import_computers.html')

    file = request.files.get('file')
    if not file or file.filename == '':
        flash('No se seleccionó ningún archivo.', 'error')
        return redirect(url_for('import_computers'))

    data = file.read()
    filename = (file.filename or '').lower()
    rows = []
    errors = []

    try:
        if filename.endswith('.csv'):
            text = data.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            for r in reader:
                rows.append(r)
        elif filename.endswith('.xlsx') or filename.endswith('.xlsm') or filename.endswith('.xls'):
            wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            try:
                headers = [str(h).strip() if h is not None else '' for h in next(it)]
            except StopIteration:
                headers = []
            for row in it:
                rowdict = {}
                for i, h in enumerate(headers):
                    key = h if h is not None else f'col{i}'
                    val = row[i] if i < len(row) else None
                    if val is not None:
                        if isinstance(val, float) and val.is_integer():
                            val = str(int(val))
                        else:
                            val = str(val)
                    rowdict[key] = val
                rows.append(rowdict)
        else:
            flash('Formato no soportado. Suba CSV o XLSX.', 'error')
            return redirect(url_for('import_computers'))
    except Exception as e:
        flash(f'Error al procesar el archivo: {e}', 'error')
        return redirect(url_for('import_computers'))

    inserted = 0
    if rows:
        db = get_db()
        for r in rows:
            proyecto = _get_value(r, ['proyecto', 'PROYECTO']) or 'Mitie'
            hostname = _get_value(r, ['hostname', 'HOSTNAME', 'equipo'])
            numero_serie = _get_value(r, ['numero_serie', 'serial', 'sn', 'SN'])
            apellidos_nombre = _get_value(r, ['apellidos_nombre', 'persona', 'usuario_equipo'])
            notas = _get_value(r, ['notas', 'observaciones'])
            tipo = _get_value(r, ['tipo', 'TIPO']) or 'Entrega'
            usuario = current_user.username
            timestamp = datetime.now().isoformat()

            if not hostname:
                errors.append(f'Fila sin hostname omitida.')
                continue

            try:
                db.execute('INSERT INTO computers (hostname, numero_serie, apellidos_nombre, notas, tipo, usuario, timestamp, proyecto) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                           (hostname, numero_serie, apellidos_nombre, notas, tipo, usuario, timestamp, proyecto))
                inserted += 1
            except Exception as e:
                errors.append(f'Error insertando equipo {hostname}: {e}')
        db.commit()

    flash(f'Se insertaron {inserted} registros de computers.', 'success' if inserted > 0 else 'warning')
    if errors:
        for err in errors[:5]:
            flash(err, 'error')
    return redirect(url_for('history_computers_entrega'))


def _get_value(row, keys):
    # Busca una lista de posibles nombres (case-insensitive) en el diccionario row
    if not row:
        return ''
    for k in keys:
        # exact key
        if k in row and row[k] is not None:
            return str(row[k]).strip()
    # try case-insensitive
    lower_map = { (str(kk).lower() if kk is not None else ''): vv for kk,vv in row.items() }
    for k in keys:
        kk = k.lower()
        if kk in lower_map and lower_map[kk] is not None:
            return str(lower_map[kk]).strip()
    return ''


@app.route('/import', methods=['GET', 'POST'])
@require_permission('registrar')
def import_file():
    if request.method == 'GET':
        return render_template('import.html')

    file = request.files.get('file')
    if not file or file.filename == '':
        return render_template('import_result.html', inserted=0, errors=['No se subió ningún archivo.'])

    data = file.read()
    filename = (file.filename or '').lower()
    rows = []
    errors = []

    try:
        if filename.endswith('.csv'):
            text = data.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            for r in reader:
                rows.append(r)
        elif filename.endswith('.xlsx') or filename.endswith('.xlsm') or filename.endswith('.xls'):
            wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            try:
                headers = [str(h).strip() if h is not None else '' for h in next(it)]
            except StopIteration:
                headers = []
            for row in it:
                rowdict = {}
                for i, h in enumerate(headers):
                    key = h if h is not None else f'col{i}'
                    val = row[i] if i < len(row) else None
                    # Convertir valores numéricos a string sin decimales
                    if val is not None:
                        if isinstance(val, float) and val.is_integer():
                            val = str(int(val))
                        else:
                            val = str(val)
                    rowdict[key] = val
                rows.append(rowdict)
        else:
            errors.append('Formato no soportado. Suba CSV o XLSX.')
    except Exception as e:
        errors.append(f'Error al procesar el archivo: {e}')

    inserted = 0
    if rows:
        db = get_db()
        for r in rows:
            situm = _get_value(r, ['situm', 'SITUM'])
            usuario = _get_value(r, ['usuario', 'user', 'nombre'])
            imei = _get_value(r, ['imei', 'IMEI'])
            raw_tel = _get_value(r, ['telefono', 'phone', 'telefono_movil']) or ''
            telefono = format_phone(raw_tel)
            if raw_tel and telefono is None:
                errors.append(f'Fila con IMEI={imei}: teléfono inválido "{raw_tel}"')
                continue
            notas_telefono = _get_value(r, ['notas_telefono', 'notas', 'notes', 'modelo', 'model'])
            tipo = _get_value(r, ['tipo', 'type']) or 'entrega'
            timestamp = datetime.utcnow().isoformat()
            try:
                db.execute('INSERT INTO entregas (situm, usuario, imei, telefono, notas_telefono, tipo, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)',
                           (situm, usuario, imei, telefono, notas_telefono, tipo, timestamp))
                inserted += 1
            except Exception as e:
                errors.append(f'Error insertando fila con IMEI={imei}: {e}')
        db.commit()

    return render_template('import_result.html', inserted=inserted, errors=errors)

@app.route('/incidents')
@require_permission('ver_incidencias')
def incidents():
    db = get_db()
    imei_search = request.args.get('imei', '').strip()
    usuario_search = request.args.get('usuario', '').strip()
    fecha_inicio = request.args.get('fecha_inicio', '').strip()
    fecha_fin = request.args.get('fecha_fin', '').strip()
    
    query = 'SELECT * FROM incidencias WHERE 1=1'
    params = []
    
    if imei_search:
        query += ' AND imei LIKE ?'
        params.append(f'%{imei_search}%')
    
    if usuario_search:
        query += ' AND usuario LIKE ?'
        params.append(f'%{usuario_search}%')
    
    if fecha_inicio:
        query += ' AND timestamp >= ?'
        params.append(f'{fecha_inicio}T00:00:00')
    
    if fecha_fin:
        query += ' AND timestamp <= ?'
        params.append(f'{fecha_fin}T23:59:59')
    
    query += ' ORDER BY timestamp DESC'
    
    cur = db.execute(query, params)
    incidents = cur.fetchall()
    return render_template('incidents.html', incidents=incidents, imei_search=imei_search, usuario_search=usuario_search, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)


@app.route('/registro/<int:registro_id>/editar', methods=['GET', 'POST'])
@require_permission('administracion')
def editar_registro(registro_id):
    db = get_db()
    cur = db.execute('SELECT * FROM entregas WHERE id = ?', (registro_id,))
    r = cur.fetchone()
    if not r:
        flash('Registro no encontrado', 'error')
        return redirect(url_for('history'))

    if request.method == 'GET':
        return render_template('edit_entrega.html', r=r)

    # POST: procesar actualización
    situm = request.form.get('situm', '').strip()
    usuario = request.form.get('usuario', '').strip()
    imei = request.form.get('imei', '').strip()
    raw_tel = request.form.get('telefono', '').strip()
    telefono = format_phone(raw_tel)
    notas_telefono = request.form.get('notas_telefono', '').strip()

    # Validaciones
    if situm and not is_mitie_email(situm):
        flash('El campo Situm debe ser un email con dominio @mitie.es', 'error')
        return redirect(url_for('editar_registro', registro_id=registro_id))

    imei_digits = re.sub(r'\D', '', imei or '')
    if imei and not is_valid_imei(imei_digits):
        flash('IMEI inválido — debe contener exactamente 15 dígitos', 'error')
        return redirect(url_for('editar_registro', registro_id=registro_id))

    if raw_tel and telefono is None:
        flash('Teléfono inválido — debe ser numérico de 9 dígitos o con prefijo +34/0034/34', 'error')
        return redirect(url_for('editar_registro', registro_id=registro_id))

    db.execute('UPDATE entregas SET situm = ?, usuario = ?, imei = ?, telefono = ?, notas_telefono = ? WHERE id = ?'
               , (situm, usuario, imei_digits, telefono, notas_telefono, registro_id))
    db.commit()
    flash('Registro actualizado correctamente', 'success')
    return redirect(url_for('history'))


@app.route('/incidencia/<int:inc_id>/editar', methods=['GET', 'POST'])
@require_permission('administracion')
def editar_incidencia(inc_id):
    db = get_db()
    cur = db.execute('SELECT * FROM incidencias WHERE id = ?', (inc_id,))
    i = cur.fetchone()
    if not i:
        flash('Incidencia no encontrada', 'error')
        return redirect(url_for('incidents'))

    if request.method == 'GET':
        return render_template('edit_incidencia.html', i=i)

    imei = request.form.get('imei', '').strip()
    usuario = request.form.get('usuario', '').strip()
    raw_tel = request.form.get('telefono', '').strip()
    telefono = format_phone(raw_tel)
    notas = request.form.get('notas', '').strip()

    imei_digits = re.sub(r'\D', '', imei or '')
    if imei and not is_valid_imei(imei_digits):
        flash('IMEI inválido — debe contener exactamente 15 dígitos', 'error')
        return redirect(url_for('editar_incidencia', inc_id=inc_id))

    if raw_tel and telefono is None:
        flash('Teléfono inválido — debe ser numérico de 9 dígitos o con prefijo +34/0034/34', 'error')
        return redirect(url_for('editar_incidencia', inc_id=inc_id))

    db.execute('UPDATE incidencias SET imei = ?, usuario = ?, telefono = ?, notas = ? WHERE id = ?',
               (imei_digits, usuario, telefono, notas, inc_id))
    db.commit()
    flash('Incidencia actualizada correctamente', 'success')
    return redirect(url_for('incidents'))

@app.route('/incidents/export')
@require_permission('ver_incidencias')
def export_incidents():
    db = get_db()
    imei_search = request.args.get('imei', '').strip()
    usuario_search = request.args.get('usuario', '').strip()
    ids_param = request.args.get('ids', '').strip()

    # If specific ids provided, export only those
    if ids_param:
        id_list = [int(i) for i in ids_param.split(',') if i.strip().isdigit()]
        if not id_list:
            rows = []
        else:
            placeholders = ','.join(['?'] * len(id_list))
            query = f'SELECT id, imei, usuario, telefono, notas, archivo_nombre, timestamp FROM incidencias WHERE id IN ({placeholders}) ORDER BY timestamp DESC'
            cur = db.execute(query, id_list)
            rows = cur.fetchall()
    else:
        query = 'SELECT id, imei, usuario, telefono, notas, archivo_nombre, timestamp FROM incidencias WHERE 1=1'
        params = []
        if imei_search:
            query += ' AND imei LIKE ?'
            params.append(f'%{imei_search}%')
        if usuario_search:
            query += ' AND usuario LIKE ?'
            params.append(f'%{usuario_search}%')
        query += ' ORDER BY timestamp DESC'

        cur = db.execute(query, params)
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.append(['IMEI', 'Usuario', 'Teléfono', 'Notas', 'Archivo', 'Fecha (UTC)'])
    for r in rows:
        ws.append([
            r['imei'] if r['imei'] else '',
            r['usuario'] if r['usuario'] else '',
            r['telefono'] if r['telefono'] else '',
            r['notas'] if r['notas'] else '',
            r['archivo_nombre'] if r['archivo_nombre'] else '',
            r['timestamp'] if r['timestamp'] else '',
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name='incidencias.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    db = get_db()

    if request.method == 'POST':
        notas = request.form.get('notas', '').strip()
        db.execute('UPDATE usuarios SET notas = ? WHERE id = ?', (notas, current_user.id))
        db.commit()
        flash('Notas actualizadas correctamente', 'success')
        return redirect(url_for('perfil'))

    usuario = db.execute(
        'SELECT id, username, rol, notas, fecha_creacion FROM usuarios WHERE id = ?',
        (current_user.id,)
    ).fetchone()
    return render_template('perfil.html', usuario=usuario)


@app.route('/perfil/cambiar_contrasena', methods=['GET', 'POST'])
@login_required
def cambiar_contrasena():
    db = get_db()
    cursor = db.cursor()
    if request.method == 'POST':
        current = request.form.get('current_password', '')
        new = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')

        if not current or not new or not confirm:
            flash('Todos los campos son requeridos', 'error')
            return redirect(url_for('cambiar_contrasena'))

        if new != confirm:
            flash('La nueva contraseña y la confirmación no coinciden', 'error')
            return redirect(url_for('cambiar_contrasena'))

        cursor.execute('SELECT password FROM usuarios WHERE id = ?', (current_user.id,))
        row = cursor.fetchone()
        if not row:
            flash('Usuario no encontrado', 'error')
            return redirect(url_for('index'))

        # row may be sqlite3.Row
        stored_hash = row['password'] if isinstance(row, dict) or hasattr(row, 'keys') else row[0]
        if not check_password_hash(stored_hash, current):
            flash('Contraseña actual incorrecta', 'error')
            return redirect(url_for('cambiar_contrasena'))

        if len(new) < 6:
            flash('La contraseña debe tener al menos 6 caracteres', 'error')
            return redirect(url_for('cambiar_contrasena'))

        new_hash = generate_password_hash(new)
        db.execute('UPDATE usuarios SET password = ? WHERE id = ?', (new_hash, current_user.id))
        db.commit()
        flash('Contraseña actualizada correctamente', 'success')
        return redirect(url_for('perfil'))

    return render_template('cambiar_contrasena.html')

@app.route('/usuarios_gtd_sgpmr')
@require_permission('registrar')
def usuarios_gtd_sgpmr():
    db = get_db()
    usuarios = db.execute('SELECT * FROM usuarios_gtd_sgpmr ORDER BY fecha_creacion DESC').fetchall()
    return render_template('usuarios_gtd_sgpmr.html', usuarios=usuarios)

@app.route('/usuarios_gtd_sgpmr/crear', methods=['GET', 'POST'])
@require_permission('registrar')
def crear_usuario_gtd_sgpmr():
    db = get_db()
    if request.method == 'POST':
        usuario_gtd = request.form.get('usuario_gtd', '').strip()
        usuario_sgpmr = request.form.get('usuario_sgpmr', '').strip()
        nombre_apellidos = request.form.get('nombre_apellidos', '').strip()
        correo_electronico = request.form.get('correo_electronico', '').strip()
        dni_nie = request.form.get('dni_nie', '').strip()
        
        if not nombre_apellidos:
            flash('El nombre y apellidos es requerido', 'error')
            return redirect(url_for('crear_usuario_gtd_sgpmr'))
        
        try:
            db.execute('''
                INSERT INTO usuarios_gtd_sgpmr (usuario_gtd, usuario_sgpmr, nombre_apellidos, correo_electronico, dni_nie, fecha_creacion)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (usuario_gtd or None, usuario_sgpmr or None, nombre_apellidos, correo_electronico or None, dni_nie or None, datetime.utcnow().isoformat()))
            db.commit()
            flash('Usuario creado correctamente', 'success')
            return redirect(url_for('usuarios_gtd_sgpmr'))
        except sqlite3.IntegrityError as e:
            flash(f'Error al crear usuario: {str(e)}', 'error')
            return redirect(url_for('crear_usuario_gtd_sgpmr'))
    
    return render_template('crear_usuario_gtd_sgpmr.html')

@app.route('/usuarios_gtd_sgpmr/<int:usuario_id>/editar', methods=['GET', 'POST'])
@require_permission('registrar')
def editar_usuario_gtd_sgpmr(usuario_id):
    db = get_db()
    usuario = db.execute('SELECT * FROM usuarios_gtd_sgpmr WHERE id = ?', (usuario_id,)).fetchone()
    
    if not usuario:
        flash('Usuario no encontrado', 'error')
        return redirect(url_for('usuarios_gtd_sgpmr'))
    
    if request.method == 'POST':
        usuario_gtd = request.form.get('usuario_gtd', '').strip()
        usuario_sgpmr = request.form.get('usuario_sgpmr', '').strip()
        nombre_apellidos = request.form.get('nombre_apellidos', '').strip()
        correo_electronico = request.form.get('correo_electronico', '').strip()
        dni_nie = request.form.get('dni_nie', '').strip()
        
        if not nombre_apellidos:
            flash('El nombre y apellidos es requerido', 'error')
            return redirect(url_for('editar_usuario_gtd_sgpmr', usuario_id=usuario_id))
        
        try:
            db.execute('''
                UPDATE usuarios_gtd_sgpmr
                SET usuario_gtd = ?, usuario_sgpmr = ?, nombre_apellidos = ?, correo_electronico = ?, dni_nie = ?
                WHERE id = ?
            ''', (usuario_gtd or None, usuario_sgpmr or None, nombre_apellidos, correo_electronico or None, dni_nie or None, usuario_id))
            db.commit()
            flash('Usuario actualizado correctamente', 'success')
            return redirect(url_for('usuarios_gtd_sgpmr'))
        except sqlite3.IntegrityError as e:
            flash(f'Error al actualizar usuario: {str(e)}', 'error')
            return redirect(url_for('editar_usuario_gtd_sgpmr', usuario_id=usuario_id))
    
    return render_template('editar_usuario_gtd_sgpmr.html', usuario=usuario)

def check_admin_password(password):
    if not password:
        return False
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT password FROM usuarios WHERE username = "admin"')
    row = cursor.fetchone()
    if row and check_password_hash(row[0], password):
        return True
    return False

@app.route('/usuarios_gtd_sgpmr/<int:usuario_id>/eliminar', methods=['POST'])
@require_permission('registrar')
def eliminar_usuario_gtd_sgpmr(usuario_id):
    password = request.form.get('admin_password')
    if not check_admin_password(password):
        flash('Contraseña de administrador incorrecta', 'error')
        return redirect(url_for('usuarios_gtd_sgpmr'))

    db = get_db()
    db.execute('DELETE FROM usuarios_gtd_sgpmr WHERE id = ?', (usuario_id,))
    db.commit()
    flash('Usuario eliminado correctamente', 'success')
    return redirect(url_for('usuarios_gtd_sgpmr'))

@app.route('/usuarios_gtd_sgpmr/importar', methods=['GET', 'POST'])
@require_permission('registrar')
def importar_usuarios_gtd_sgpmr():
    db = get_db()
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionó archivo', 'error')
            return redirect(url_for('importar_usuarios_gtd_sgpmr'))
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            flash('No se seleccionó archivo', 'error')
            return redirect(url_for('importar_usuarios_gtd_sgpmr'))
        
        try:
            if archivo.filename.endswith('.csv'):
                stream = io.StringIO(archivo.stream.read().decode('UTF-8'))
                csv_data = list(csv.DictReader(stream))
            elif archivo.filename.endswith('.xlsx'):
                wb = load_workbook(archivo.stream)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                csv_data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    csv_data.append(dict(zip(headers, row)))
            else:
                flash('Formato no soportado. Use CSV o XLSX', 'error')
                return redirect(url_for('importar_usuarios_gtd_sgpmr'))
            
            inserted = 0
            errors = []
            for idx, row in enumerate(csv_data, 2):
                try:
                    usuario_gtd = str(row.get('usuario_gtd') or '').strip() or None
                    usuario_sgpmr = str(row.get('usuario_sgpmr') or '').strip() or None
                    nombre_apellidos = str(row.get('nombre_apellidos') or '').strip()
                    correo_electronico = str(row.get('correo_electronico') or '').strip() or None
                    dni_nie = str(row.get('dni_nie') or '').strip() or None
                    
                    if not nombre_apellidos:
                        errors.append(f"Fila {idx}: nombre_apellidos es requerido")
                        continue
                    
                    db.execute('''
                        INSERT INTO usuarios_gtd_sgpmr (usuario_gtd, usuario_sgpmr, nombre_apellidos, correo_electronico, dni_nie, fecha_creacion)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (usuario_gtd, usuario_sgpmr, nombre_apellidos, correo_electronico, dni_nie, datetime.utcnow().isoformat()))
                    inserted += 1
                except sqlite3.IntegrityError as e:
                    errors.append(f"Fila {idx}: {str(e)}")
            
            db.commit()
            flash(f'Se importaron {inserted} usuarios correctamente. Errores: {len(errors)}', 'success' if inserted > 0 else 'warning')
            if errors:
                for error in errors[:10]:
                    flash(error, 'error')
            return redirect(url_for('usuarios_gtd_sgpmr'))
        except Exception as e:
            flash(f'Error al procesar archivo: {str(e)}', 'error')
            return redirect(url_for('importar_usuarios_gtd_sgpmr'))
    
    return render_template('importar_usuarios_gtd_sgpmr.html')

@app.route('/incidents/delete-selected', methods=['POST'])
@require_permission('borrar_registros')
def delete_selected_incidents():
    """Borrar incidencias seleccionadas con verificación de contraseña"""
    password = request.form.get('password', '').strip()
    ids_param = request.form.get('ids', '').strip()
    
    # Verificar contraseña
    if password != '4j_6WbTT7scyicJcam':
        return redirect(url_for('incidents') + '?error=Invalid password')
    
    if ids_param:
        id_list = [int(i) for i in ids_param.split(',') if i.strip().isdigit()]
        if id_list:
            db = get_db()
            placeholders = ','.join(['?'] * len(id_list))
            db.execute(f'DELETE FROM incidencias WHERE id IN ({placeholders})', id_list)
            db.commit()
    
    return redirect(url_for('incidents'))

@app.route('/incidents/download/<int:incident_id>')
@require_permission('ver_incidencias')
def download_incident_file(incident_id):
    db = get_db()
    db.row_factory = sqlite3.Row
    incident = db.execute('SELECT * FROM incidencias WHERE id = ?', (incident_id,)).fetchone()
    
    if not incident:
        return "Incidencia no encontrada", 404
    
    return send_file(
        io.BytesIO(incident['archivo_contenido']),
        as_attachment=True,
        download_name=incident['archivo_nombre']
    )

@app.route('/inventario_telefonos')
@require_permission('registrar')
def inventario_telefonos():
    db = get_db()
    telefonos = db.execute('SELECT * FROM inventario_telefonos ORDER BY fecha_creacion DESC').fetchall()
    return render_template('inventario_telefonos.html', telefonos=telefonos)

@app.route('/inventario_telefonos/crear', methods=['GET', 'POST'])
@require_permission('registrar')
def crear_inventario_telefonos():
    db = get_db()
    if request.method == 'POST':
        imei = request.form.get('imei', '').strip()
        numero_serie = request.form.get('numero_serie', '').strip()
        modelo = request.form.get('modelo', '').strip()
        telefono_asociado = request.form.get('telefono_asociado', '').strip()
        
        if not imei:
            flash('El IMEI es requerido', 'error')
            return redirect(url_for('crear_inventario_telefonos'))
        
        try:
            db.execute('''
                INSERT INTO inventario_telefonos (imei, numero_serie, modelo, telefono_asociado, fecha_creacion)
                VALUES (?, ?, ?, ?, ?)
            ''', (imei, numero_serie or None, modelo or None, telefono_asociado or None, datetime.utcnow().isoformat()))
            db.commit()
            flash('Teléfono registrado correctamente', 'success')
            return redirect(url_for('inventario_telefonos'))
        except sqlite3.IntegrityError as e:
            flash(f'Error al registrar teléfono: {str(e)}', 'error')
            return redirect(url_for('crear_inventario_telefonos'))
    
    return render_template('crear_inventario_telefonos.html')

@app.route('/inventario_telefonos/<int:telefono_id>/editar', methods=['GET', 'POST'])
@require_permission('registrar')
def editar_inventario_telefonos(telefono_id):
    db = get_db()
    telefono = db.execute('SELECT * FROM inventario_telefonos WHERE id = ?', (telefono_id,)).fetchone()
    
    if not telefono:
        flash('Teléfono no encontrado', 'error')
        return redirect(url_for('inventario_telefonos'))
    
    if request.method == 'POST':
        imei = request.form.get('imei', '').strip()
        numero_serie = request.form.get('numero_serie', '').strip()
        modelo = request.form.get('modelo', '').strip()
        telefono_asociado = request.form.get('telefono_asociado', '').strip()
        
        if not imei:
            flash('El IMEI es requerido', 'error')
            return redirect(url_for('editar_inventario_telefonos', telefono_id=telefono_id))
        
        try:
            db.execute('''
                UPDATE inventario_telefonos
                SET imei = ?, numero_serie = ?, modelo = ?, telefono_asociado = ?
                WHERE id = ?
            ''', (imei, numero_serie or None, modelo or None, telefono_asociado or None, telefono_id))
            db.commit()
            flash('Teléfono actualizado correctamente', 'success')
            return redirect(url_for('inventario_telefonos'))
        except sqlite3.IntegrityError as e:
            flash(f'Error al actualizar teléfono: {str(e)}', 'error')
            return redirect(url_for('editar_inventario_telefonos', telefono_id=telefono_id))
    
    return render_template('editar_inventario_telefonos.html', telefono=telefono)

@app.route('/inventario_telefonos/<int:telefono_id>/eliminar', methods=['POST'])
@require_permission('registrar')
def eliminar_inventario_telefonos(telefono_id):
    password = request.form.get('admin_password')
    if not check_admin_password(password):
        flash('Contraseña de administrador incorrecta', 'error')
        return redirect(url_for('inventario_telefonos'))

    db = get_db()
    db.execute('DELETE FROM inventario_telefonos WHERE id = ?', (telefono_id,))
    db.commit()
    flash('Teléfono eliminado correctamente', 'success')
    return redirect(url_for('inventario_telefonos'))

@app.route('/inventario_telefonos/delete-selected', methods=['POST'])
@require_permission('registrar')
def delete_selected_inventario_telefonos():
    ids_param = request.form.get('ids', '').strip()
    password = request.form.get('admin_password')
    
    if not check_admin_password(password):
        flash('Contraseña de administrador incorrecta', 'error')
        return redirect(url_for('inventario_telefonos'))

    if ids_param:
        id_list = [int(i) for i in ids_param.split(',') if i.strip().isdigit()]
        if id_list:
            db = get_db()
            placeholders = ','.join(['?'] * len(id_list))
            db.execute(f'DELETE FROM inventario_telefonos WHERE id IN ({placeholders})', id_list)
            db.commit()
            flash('Teléfonos eliminados correctamente', 'success')

    return redirect(url_for('inventario_telefonos'))

@app.route('/inventario_telefonos/importar', methods=['GET', 'POST'])
@require_permission('registrar')
def importar_inventario_telefonos():
    db = get_db()
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionó archivo', 'error')
            return redirect(url_for('importar_inventario_telefonos'))
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            flash('No se seleccionó archivo', 'error')
            return redirect(url_for('importar_inventario_telefonos'))
        
        try:
            if archivo.filename.endswith('.csv'):
                stream = io.StringIO(archivo.stream.read().decode('UTF-8'))
                csv_data = list(csv.DictReader(stream))
            elif archivo.filename.endswith('.xlsx'):
                wb = load_workbook(archivo.stream)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                csv_data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    csv_data.append(dict(zip(headers, row)))
            else:
                flash('Formato no soportado. Use CSV o XLSX', 'error')
                return redirect(url_for('importar_inventario_telefonos'))
            
            inserted = 0
            errors = []
            for idx, row in enumerate(csv_data, 2):
                try:
                    imei = str(row.get('imei') or '').strip()
                    numero_serie = str(row.get('numero_serie') or '').strip() or None
                    modelo = str(row.get('modelo') or '').strip() or None
                    telefono_asociado = str(row.get('telefono_asociado') or '').strip() or None
                    
                    if not imei:
                        errors.append(f"Fila {idx}: IMEI es requerido")
                        continue
                    
                    db.execute('''
                        INSERT INTO inventario_telefonos (imei, numero_serie, modelo, telefono_asociado, fecha_creacion)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (imei, numero_serie, modelo, telefono_asociado, datetime.utcnow().isoformat()))
                    inserted += 1
                except sqlite3.IntegrityError as e:
                    errors.append(f"Fila {idx}: {str(e)}")
            
            db.commit()
            flash(f'Se importaron {inserted} teléfonos correctamente. Errores: {len(errors)}', 'success' if inserted > 0 else 'warning')
            if errors:
                for error in errors[:10]:
                    flash(error, 'error')
            return redirect(url_for('inventario_telefonos'))
        except Exception as e:
            flash(f'Error al procesar archivo: {str(e)}', 'error')
            return redirect(url_for('importar_inventario_telefonos'))
    
    return render_template('importar_inventario_telefonos.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
