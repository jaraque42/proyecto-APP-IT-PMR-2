"""Funciones de utilidad compartidas: validación, email, PDF, importación y paginación."""

import csv
import io
import math
import os
import random
import re
import smtplib
import ssl
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from flask import request
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Validación
# ---------------------------------------------------------------------------

def format_phone(phone):
    """Normaliza teléfono: devuelve solo dígitos (9 dígitos nacionales).

    Retorna '' si vacío, None si inválido, o el teléfono normalizado.
    """
    if not phone:
        return ''
    phone = str(phone).strip()
    digits = re.sub(r'\D', '', phone)
    if not digits:
        return None
    if len(digits) == 9:
        return digits
    if digits.startswith('34') and len(digits) == 11:
        return digits[2:]
    if digits.startswith('0034') and len(digits) == 13:
        return digits[4:]
    if digits.startswith('0') and len(digits) == 10:
        return digits[1:]
    return None


def is_mitie_email(addr: str) -> bool:
    if not addr:
        return False
    return bool(re.match(r'^[A-Za-z0-9._%+-]+@mitie\.es$', addr.strip(), re.IGNORECASE))


def is_valid_imei(imei: str) -> bool:
    if not imei:
        return False
    return bool(re.match(r'^\d{15}$', imei.strip()))


# ---------------------------------------------------------------------------
# Email / OTP
# ---------------------------------------------------------------------------

def send_validation_email_verbose(to_email, codigo):
    """Envía correo de validación. Devuelve dict con 'success' y opcionalmente 'error'."""
    smtp_server = os.environ.get('SMTP_SERVER', 'localhost')
    smtp_port = int(os.environ.get('SMTP_PORT', '465'))
    smtp_user = os.environ.get('SMTP_USER', '')
    smtp_pass = os.environ.get('SMTP_PASS', '')
    smtp_from = os.environ.get('SMTP_FROM', smtp_user)

    if not smtp_user:
        return {'success': True}  # modo debug sin SMTP configurado

    try:
        msg = MIMEMultipart()
        msg['From'] = smtp_from
        msg['To'] = to_email
        msg['Subject'] = f'Código de Validación: {codigo} - Mitie'
        msg.attach(MIMEText(f'Tu código de validación para la entrega de dispositivo es: {codigo}', 'plain'))

        try:
            context = ssl.create_default_context()
        except Exception:
            context = ssl._create_unverified_context()

        if smtp_port == 465:
            server = smtplib.SMTP_SSL(smtp_server, smtp_port, context=context, timeout=15)
        else:
            server = smtplib.SMTP(smtp_server, smtp_port, timeout=15)
            server.starttls(context=context)

        with server:
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def send_validation_email(to_email, codigo):
    """Wrapper de compatibilidad. Devuelve bool."""
    return send_validation_email_verbose(to_email, codigo)['success']


# ---------------------------------------------------------------------------
# Generación de PDF de entrega
# ---------------------------------------------------------------------------

_TEXTO_COMUNICACION = """Mitie Facilities Services, S.A. C/ Juan Ignacio Luca de Tena, 8 - 1° 28027 Madrid España

COMUNICACIÓN

Para mejorar la organización del servicio, evitar incidentes con los materiales de la Empresa, conseguir los estándares de calidad y mejorar la productividad que requiere nuestro servicio, a través de la entrega del dispositivo móvil (PDA) propiedad de la Empresa y de los equipos de trabajo que pone la Empresa a su disposición, se inicia la implantación de un sistema de geolocalización de estos dispositivos (PDA) y herramientas como sillas de ruedas cuyo uso se encuentra limitado a los fines anteriores y exclusivamente durante la prestación de SU jornada de trabajo en el Centro del Aeropuerto.

Para su conocimiento, se ha comunicado con anterioridad a la Representación Legal de las Personas de este Servicio PMR del Centro del Aeropuerto y con carácter previo a su implantación a través de diversas comunicaciones, principalmente efectuadas el 26 de septiembre, 11 de octubre y el pasado 28 de octubre del año en curso, con información detallada del sistema.

Le recordamos que el uso de los equipos y particularmente dispositivos móviles (PDA) entregados por la Empresa están destinados solo y exclusivamente para desempeño de su labor en el Servicio en el Centro del Aeropuerto, quedando el uso exclusivamente limitado a la prestación de sus servicios dentro de su jornada de trabajo, no estando permitido su utilización fuera de esta.

Por consiguiente:

1. La implantación no comporta modificación de sistema organizativo.

2. Se utiliza en herramientas y dispositivos móviles propiedad de la Empresa, protegiéndose la intimidad de todas las personas trabajadoras del servicio. Se trata por tanto de una geolocalización admitida, sobre Medios propiedad de la Empresa acordes al art. 20 bis ET.

3. Los datos relativos al posicionamiento geográfico que proporciona el uso del sistema de geolocalización, no constituyen datos de carácter personal salvo los ya recogido en la entrega del equipo, ya que se centran en medios y herramientas de la Empresa, si bien, garantizamos siempre y en todo caso que estos datos serán tratados de conformidad con lo dispuesto en el Reglamento (UE) 2016/679, de 27 de abril (GDPR), y la Ley Orgánica 3/2018, de 5 de diciembre (LOPDGDD), con el apoyo del Delegado de Protección de Datos (dpd@acoran.es D. Mario García.).

4. Este sistema cumple con todos los requisitos a nivel legal, respetando en todo momento los principios de proporcionalidad, necesidad e idoneidad.

5. La Geolocalización, le recordamos, solo se realiza cuándo el equipo está operativo, durante su jornada de trabajo y en las zonas públicas del entorno Aeroportuario, respetando los periodos de descanso como refrigerio.
"""


def generate_entrega_pdf(situm, usuario, imei, telefono, notas, timestamp, codigo_validacion=None):
    """Genera PDF de entrega y devuelve (buffer, filename)."""
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                 fontSize=14, textColor=colors.HexColor('#333333'),
                                 spaceAfter=12, alignment=1)
    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'],
                                  fontSize=9, leading=11, spaceAfter=6)

    elements = []

    # Logo
    logo_path = os.path.join(BASE_DIR, 'static', 'mitie_logo.png')
    if os.path.exists(logo_path):
        try:
            img_reader = ImageReader(logo_path)
            img_w, img_h = img_reader.getSize()
            desired_width = 2 * inch
            desired_height = desired_width * (float(img_h) / float(img_w)) if img_w else desired_width * 0.5
            page_height = doc.pagesize[1]
            usable_height = page_height - doc.topMargin - doc.bottomMargin
            max_height = usable_height * 0.25
            if desired_height > max_height:
                scale = max_height / desired_height
                desired_width *= scale
                desired_height = max_height
            logo = Image(logo_path, width=desired_width, height=desired_height)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.15 * inch))
        except Exception:
            pass

    elements.append(Paragraph("REGISTRO DE ENTREGA DIGITAL", title_style))
    elements.append(Spacer(1, 0.2 * inch))

    data = [
        ['CAMPO', 'VALOR'],
        ['Situm', situm or ''],
        ['Usuario', usuario or ''],
        ['IMEI', imei or ''],
        ['Teléfono', telefono or ''],
        ['Notas', notas or ''],
        ['Fecha', timestamp[:10] if timestamp else ''],
    ]
    table = Table(data, colWidths=[1.5 * inch, 4 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.HexColor('#ffffff')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.3 * inch))

    elements.append(Paragraph("COMUNICACIÓN SOBRE SISTEMA DE GEOLOCALIZACIÓN", title_style))
    elements.append(Spacer(1, 0.1 * inch))
    elements.append(Paragraph(_TEXTO_COMUNICACION, normal_style))
    elements.append(Spacer(1, 0.3 * inch))

    if codigo_validacion:
        elements.append(Paragraph(f"<b>FIRMA DIGITAL (Validada por Email):</b> {codigo_validacion}", normal_style))
        elements.append(Paragraph(f"Confirmado electrónicamente el {timestamp[:10]}", normal_style))
        elements.append(Spacer(1, 0.2 * inch))

    doc.build(elements)
    pdf_buffer.seek(0)

    # Guardar copia en servidor
    pdf_dir = os.path.join(BASE_DIR, 'pdfs', 'entregas')
    os.makedirs(pdf_dir, exist_ok=True)
    pdf_filename = f"entrega_{imei}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf_path = os.path.join(pdf_dir, pdf_filename)
    with open(pdf_path, 'wb') as f:
        f.write(pdf_buffer.getvalue())

    pdf_buffer.seek(0)
    return pdf_buffer, pdf_filename


# ---------------------------------------------------------------------------
# Importación genérica de ficheros CSV / XLSX
# ---------------------------------------------------------------------------

def parse_import_file(file):
    """Lee un fichero CSV o XLSX y devuelve (rows: list[dict], errors: list[str]).

    ``file`` es un FileStorage de Flask (request.files).
    """
    if not file or file.filename == '':
        return [], ['No se subió ningún archivo.']

    data = file.read()
    filename = (file.filename or '').lower()
    rows = []
    errors = []

    try:
        if filename.endswith('.csv'):
            text = data.decode('utf-8-sig')
            rows = list(csv.DictReader(io.StringIO(text)))
        elif filename.endswith(('.xlsx', '.xlsm', '.xls')):
            wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            try:
                headers = [str(h).strip() if h is not None else '' for h in next(it)]
            except StopIteration:
                headers = []
            for row in it:
                rowdict = {}
                for i, h in enumerate(headers):
                    key = h if h else f'col{i}'
                    val = row[i] if i < len(row) else None
                    if val is not None:
                        if isinstance(val, float) and val.is_integer():
                            val = str(int(val))
                        else:
                            val = str(val)
                    rowdict[key] = val
                rows.append(rowdict)
        else:
            errors.append('Formato no soportado. Suba CSV o XLSX.')
    except Exception as e:
        errors.append(f'Error al procesar el archivo: {e}')

    return rows, errors


def get_value(row, keys):
    """Busca un valor en *row* probando varias claves (case-insensitive)."""
    if not row:
        return ''
    for k in keys:
        if k in row and row[k] is not None:
            return str(row[k]).strip()
    lower_map = {(str(kk).lower() if kk else ''): vv for kk, vv in row.items()}
    for k in keys:
        kk = k.lower()
        if kk in lower_map and lower_map[kk] is not None:
            return str(lower_map[kk]).strip()
    return ''


# ---------------------------------------------------------------------------
# Paginación
# ---------------------------------------------------------------------------

PER_PAGE = 50  # registros por página por defecto


def paginate_query(db, query, params, per_page=PER_PAGE):
    """Ejecuta *query* con paginación.

    Devuelve un dict con:
      - rows:        lista de filas de la página actual
      - page:        página actual (1-based)
      - per_page:    registros por página
      - total:       total de registros
      - total_pages: total de páginas
    """
    page = max(1, request.args.get('page', 1, type=int))

    # Total de registros
    count_query = f'SELECT COUNT(*) FROM ({query})'
    total = db.execute(count_query, params).fetchone()[0]
    total_pages = max(1, math.ceil(total / per_page))

    # Ajustar página si sobrepasa
    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page
    paginated_query = f'{query} LIMIT ? OFFSET ?'
    rows = db.execute(paginated_query, params + [per_page, offset]).fetchall()

    return {
        'rows': rows,
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': total_pages,
    }


def build_excel(headers, rows_data):
    """Crea un XLSX en memoria y devuelve un BytesIO listo para send_file."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows_data:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ---------------------------------------------------------------------------
# Verificación de contraseña de borrado
# ---------------------------------------------------------------------------

def verify_delete_password(password):
    """Compara la contraseña proporcionada con DELETE_MASTER_PASSWORD."""
    master = os.environ.get('DELETE_MASTER_PASSWORD', '')
    if not master or password != master:
        return False
    return True


def check_admin_password(password):
    """Verifica la contraseña contra el hash del usuario 'admin'."""
    if not password:
        return False
    from models import get_db
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT password FROM usuarios WHERE username = "admin"')
    row = cursor.fetchone()
    if row:
        from werkzeug.security import check_password_hash
        return check_password_hash(row[0], password)
    return False
